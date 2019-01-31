"""
"""
from itertools import product
import numpy as np
import tensorflow as tf

from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler


def load_data(data_path):
    """Load training data file."""
    # load data from file
    workbook = load_workbook(data_path)
    sheet = workbook.active
    x = []
    y = []
    precision = 1
    for row in sheet.iter_rows():
        cell_value = tuple(float(cell.value) for cell in row)
        x.append(cell_value[:-1])
        y.append([round(cell_value[-1], precision)])

    return x, y


def gen_data(x_train):
    """
    """
    layer_thickness = list(map(float, range(20, 81, 5)))
    hatch_distance = list(map(float, range(30, 81, 5)))
    laser_power = list(map(float, range(80, 181, 5)))
    laser_velocity  = list(map(float, range(800, 2501, 100)))

    x_generated = list(product(layer_thickness, hatch_distance,
                               laser_power, laser_velocity))

    x_generated = [row for row in x_generated if row not in x_train]
    x_generated = np.asarray(x_generated, dtype=np.float32)
    return x_generated


if __name__ == '__main__':
    data_path = 'cleaned_data.xlsx'
    checkpoint_dir = 'checkpoints'
    result_file = 'indexing.csv'

    # generate data
    x_train, y_train = load_data(data_path)
    x_gen = gen_data(x_train)
    print('Generated %d data points' % x_gen.shape[0])

    # normalize data
    scaler = StandardScaler()
    x = scaler.fit_transform(x_gen)

    # create a saver to load graph
    with tf.Session() as sess:
        latest_checkpoint_path = tf.train.latest_checkpoint(checkpoint_dir)
        saver = tf.train.import_meta_graph(latest_checkpoint_path + '.meta')
        saver.restore(sess, latest_checkpoint_path)

        inputs_placeholder = tf.get_default_graph().get_tensor_by_name('inputs:0')
        predictions_tensor = tf.get_default_graph().get_tensor_by_name('output:0')
        training_placeholder = tf.get_default_graph().get_tensor_by_name('training:0')
        y_preds = sess.run(predictions_tensor, feed_dict={inputs_placeholder: x, training_placeholder: False})

    # write to csv
    precision = 1
    with open(result_file, 'wt') as f:
        x_out = np.concatenate((x_gen, x_train))
        y_out = np.concatenate((y_preds, y_train))
        data = np.concatenate((x_out, y_out), axis=1)
        for row in data:
            row = np.round(row, precision)
            f.write(','.join(map(str, row)) + '\n')
    print('Created indexing file has %d data points' % x_out.shape[0])
