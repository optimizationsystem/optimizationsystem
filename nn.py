"""Train our neural networks model for slm optimizer system."""
import os
import sys
import glob
from itertools import product
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import KFold


def visualize(hist):
    """Visualize training history
    """
    plt.subplot(1, 2, 2)
    plt.plot(hist['mae'])
    plt.plot(hist['val_mae'])
    plt.title('Model MAE')
    plt.ylabel('MAE')
    plt.xlabel('epoch')
    plt.yticks([0.5, 1.0, 1.5, 2.0, 3.0], ['0.5', '1.0', '1.5', '2.0', '3.0'])
    plt.legend(['train', 'val'], loc='upper left')
    plt.savefig('train_val.png')


def gen_data(x):
    """Generate data space which excludes training data.
    """
    layer_thickness = list(map(float, range(20, 81, 5)))
    hatch_distance = list(map(float, range(30, 81, 5)))
    laser_power = list(map(float, range(80, 181, 5)))
    laser_velocity  = list(map(float, range(800, 2501, 100)))

    x_generated = list(product(layer_thickness, hatch_distance,
                               laser_power, laser_velocity))

    x_generated = [row for row in x_generated if row not in x]
    x_generated = np.asarray(x_generated, dtype=np.float32)
    return x_generated


def test_user_inputs(data_filename, checkpoint_dir, output_filename):
    """Test the model on user's inputs data.
    """
    # load raw data from file.
    workbook = load_workbook(data_filename)
    sheet = workbook.active
    precision = 1
    x = []
    y = []
    for row in sheet.iter_rows():
        cell_value = tuple(float(cell.value) for cell in row)
        x.append(cell_value[:-1])
        y.append([round(cell_value[-1], precision)])

    # predict all possible data
    x_gen = gen_data(x)
    print('Generated {} data points'.format(x_gen.shape[0]))

    # normalize data
    scaler = StandardScaler()
    norm_x_gen = scaler.fit_transform(x_gen)

    # create a saver to load graph
    with tf.Session() as sess:
        latest_checkpoint_path = tf.train.latest_checkpoint(checkpoint_dir)
        saver = tf.train.import_meta_graph(latest_checkpoint_path + '.meta')
        saver.restore(sess, latest_checkpoint_path)

        inputs_placeholder = tf.get_default_graph().get_tensor_by_name('inputs:0')
        predictions_tensor = tf.get_default_graph().get_tensor_by_name('output:0')
        training_placeholder = tf.get_default_graph().get_tensor_by_name('training:0')
        y_preds = sess.run(predictions_tensor, feed_dict={inputs_placeholder: norm_x_gen, training_placeholder: False})

    # create indexing table
    x_out = np.concatenate((x, x_gen))
    y_out = np.concatenate((y, y_preds))
    indexing_table = np.concatenate((x_out, y_out), axis=1)
    print('Created indexing table has {} points'.format(indexing_table.shape[0]))
    print('Start to predict user inputs')
    print('-' * 20)
    print('')

    # create mapping 1-value input -> 4-values output
    indexing = dict()
    for row in indexing_table:
        row = list(map(float, row))
        key = row[-1]
        key = round(key, precision)
        if key not in indexing:
            indexing[key] = []

        indexing[key].append(row[:-1])

    # test the model for user's inputs
    output_data = []
    while True:
        try:
            input_ = input('Enter your input (q/Q to quit): ')
            if isinstance(input_, str) and input_.lower() == 'q':
                print('Bye')
                break

            input_ = round(float(input_), precision)
        except Exception:
            print('Expect input as float\n')
            continue

        # check range of input
        if input_ < 75. or input_ > 100.:
            print('Input must be in range [75., 100.]\n')
            continue

        keys = np.asarray(list(indexing.keys()))
        idx = np.argmin(np.abs(keys - input_))
        min_key = round(float(keys[idx]), precision)

        print('Filter #1')
        for value in indexing[min_key]:
            print('Input: %.2f -> Output: %.2f %.2f %.2f %.2f' % (input_, *value))

        print('-' * 20)
        print('')
        print('Filter #2')
        max_idx = 0
        max_product = 0.
        for i, value in enumerate(indexing[min_key]):
            product = value[0] * value[1] * value[-1]
            if product > max_product:
                max_idx = i
                max_product = product

        print('Input: %.2f -> Output: %.2f %.2f %.2f %.2f' % (input_, *indexing[min_key][max_idx]))
        print('')
        output_data.append([input_, *indexing[min_key][max_idx]])

    # write test output data to file
    with open(output_filename, 'wt') as f:
        for row in output_data:
            f.write(','.join(map(str, row)) + '\n')
    print('Saved output data as {}'.format(output_filename))


def load_data(data_filename, test_size=20):
    """Load data from file.
    """
    workbook = load_workbook(data_filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows():
        cell_value = [float(cell.value) for cell in row]
        data.append(cell_value)

    data = np.asarray(data, dtype=np.float32)
    x, y = data[:, :-1], data[:, -1:]

    test_size = 100
    idx = x.shape[0] - test_size
    x_train, y_train = x[:idx], y[:idx]
    x_test, y_test = x[idx:], y[idx:]

    "normalized"
    scaler = StandardScaler()
    scaler.fit(x_train)
    x_train = scaler.transform(x_train)
    x_test = scaler.transform(x_test)
    return x_train, y_train, x_test, y_test


def forward(x):
    """Forward pass inputs `x` features over the networks."""
    x = tf.layers.dense(inputs=x, units=128, activation=tf.nn.relu,
                        kernel_regularizer=tf.contrib.layers.l2_regularizer(weight_decay))
    x = tf.layers.dense(inputs=x, units=128, activation=tf.nn.relu,
                        kernel_regularizer=tf.contrib.layers.l2_regularizer(weight_decay))
    x = tf.layers.dropout(inputs=x, rate=0.7, training=training)
    x = tf.layers.dense(inputs=x, units=1, activation=lambda x: tf.nn.sigmoid(x) * 25. + 75.)
    x = tf.identity(x, name='output')
    return x


def loss(labels, predictions):
    """Calculate mean squared error loss."""
    l2_loss = tf.add_n(tf.get_collection(tf.GraphKeys.REGULARIZATION_LOSSES))
    return tf.losses.mean_squared_error(labels=labels, predictions=predictions) + l2_loss


def mae(labels, predictions):
    """Calculate mean absolute error metric."""
    return tf.reduce_mean(tf.abs(labels - predictions))


def get_batch(x, y, batch_size=32):
    """Get a batch of data at index `idx`."""
    indices = np.arange(x.shape[0])
    idx = np.random.choice(indices, batch_size)
    return (x[idx], y[idx])


if __name__ == '__main__':
    data_filename = 'cleaned_data.xlsx'
    predicted_output_filename = 'predicted_user_inputs.csv'
    graph_dir = 'graph'
    checkpoint_dir = 'checkpoints'

    epochs = 300
    batch_size = 32
    learning_rate = 1e-3
    weight_decay = 5e-2

    "load data"
    x, y, x_test, y_test = load_data(data_filename)

    # define placeholders
    x_placeholder = tf.placeholder(tf.float32, [None, 4], name='inputs')
    y_placeholder = tf.placeholder(tf.float32, [None, 1], name='labels')
    training = tf.placeholder(tf.bool, name='training')

    predictions = forward(x_placeholder)
    mse_loss = loss(y_placeholder, predictions)
    mae_metric = mae(y_placeholder, predictions)

    # optimizer
    global_step = tf.Variable(0, name='global_step', trainable=False)
    optimizer = tf.train.AdamOptimizer(learning_rate)
    train_step = optimizer.minimize(mse_loss, global_step=global_step)

    with tf.Session() as sess:
        tf.global_variables_initializer().run()

        # remove latest graph for preventing to overwrite graph
        for path in glob.glob(os.path.join(graph_dir, '*')):
            os.remove(path)
        tf.summary.FileWriter(graph_dir, sess.graph)

        # create saver to store checkpoint during training
        if not os.path.exists(checkpoint_dir):
            os.makedirs(checkpoint_dir)
        checkpoint_prefix = os.path.join(checkpoint_dir, 'model')
        saver = tf.train.Saver(tf.global_variables(), max_to_keep=1)

        # 10-fold cross-validation
        skf = KFold(n_splits=10, shuffle=True, random_state=42)
        cv_mae = []
        hist = {
            'loss': [],
            'mae': [],
            'val_loss': [],
            'val_mae': []
        }
        for train_index, val_index in skf.split(x):
            x_train, y_train = x[train_index], y[train_index]
            x_val, y_val = x[val_index], y[val_index]

            for epoch in range(epochs):
                epoch_loss = 0.
                epoch_mae = 0.
                steps_per_epoch = x_train.shape[0] // batch_size + int(x_train.shape[0] % batch_size > 0)
                for i in range(steps_per_epoch):
                    x_batch, y_batch = get_batch(x_train, y_train, batch_size=batch_size)
                    _, step, running_loss, running_mae = sess.run([train_step, global_step, mse_loss, mae_metric],
                                feed_dict={x_placeholder: x_batch, y_placeholder: y_batch, training: True})

                    print('step {:04d} loss {:.4f} mae {:.4f}'.format(step, running_loss, running_mae))

                    epoch_loss += running_loss
                    epoch_mae += running_mae
                hist['loss'].append(epoch_loss / steps_per_epoch)
                hist['mae'].append(epoch_mae / steps_per_epoch)
                epoch_val_loss, epoch_val_mae = sess.run([mse_loss, mae_metric], feed_dict={x_placeholder: x_val, y_placeholder: y_val, training: False})
                hist['val_loss'].append(epoch_val_loss)
                hist['val_mae'].append(epoch_val_mae)

            val_loss, val_mae = sess.run([mse_loss, mae_metric], feed_dict={x_placeholder: x_val, y_placeholder: y_val, training: False})
            cv_mae.append(val_mae)
            print('\nevaluate loss {:.4f} mae {:.4f}'.format(val_loss, val_mae))

            # save the model
            path = saver.save(sess, checkpoint_prefix, global_step)
            print('Saved model as {}'.format(path))

        # aggregate the results
        print('average mae: {:.4f} stddev mae: {:.4f}'.format(np.mean(cv_mae), np.std(cv_mae)))

        # visualize
        visualize(hist)

        # test
        test_loss, test_mae = sess.run([mse_loss, mae_metric], feed_dict={x_placeholder: x_test, y_placeholder: y_test, training: False})
        print('\nTest loss {:.4f} mae: {:.4f}'.format(test_loss, test_mae))
        print('')
